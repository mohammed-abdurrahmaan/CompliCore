"""
Enrollment Workflow Routes - 5-Step ACA Compliance Pipeline
Step 1: Plan Library Setup (HR Admin)
Step 2: Auto-Eligibility Engine
Step 3: Employee Self-Service Portal
Step 4: HR Compliance Review
Step 5: Carrier Census Export
"""

from fastapi import HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timezone
import uuid
import csv
import io
import random

from openpyxl import Workbook
from services.mv_calculator import calculate_mv_percentage as hhs_mv_calculate
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


async def _create_notification(db, user_id, title, message, category="system", link=""):
    """Helper to create an in-app notification."""
    notif = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "title": title,
        "message": message,
        "category": category,
        "link": link,
        "read": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.notifications.insert_one(notif)
    return notif


async def _recalculate_offer_code(db, employer_id, employee_id, emp):
    """Recalculate IRS offer code based on what coverage is OFFERED to the employee.
    The offer code reflects the broadest coverage available, regardless of enrollment status."""
    is_ft = emp.get("is_full_time", False)
    if not is_ft:
        await db.eligibility_results.update_one(
            {"employer_id": employer_id, "employee_id": employee_id},
            {"$set": {"offer_code": "1H"}},
        )
        return "1H"

    # Check what plans are available to this employee
    # First: check plan_assignments (explicit HR assignments)
    assignments = await db.plan_assignments.find(
        {"employer_id": employer_id, "employee_id": employee_id}, {"_id": 0}
    ).to_list(20)
    assigned_plan_ids = [a["plan_id"] for a in assignments]

    # Also check if there's an enrollment (employee chose a plan)
    enrollment = await db.enrollments.find_one(
        {"employer_id": employer_id, "employee_id": employee_id}, {"_id": 0}
    )
    if enrollment and enrollment.get("plan_id") and enrollment["plan_id"] not in assigned_plan_ids:
        assigned_plan_ids.append(enrollment["plan_id"])

    # No explicit plan assignment or enrollment = no offer made
    if not assigned_plan_ids:
        offer_code = "1H"
    else:
        medical_plans = []
        for pid in assigned_plan_ids:
            p = await db.plan_library.find_one({"id": pid, "category": "medical"}, {"_id": 0})
            if p:
                medical_plans.append(p)

        if not medical_plans:
            offer_code = "1H"
        else:
            # Use the best (broadest) plan to determine the offer code
            primary = medical_plans[0]
            is_mec = primary.get("mec_qualified", False)
            mv_pct = primary.get("mv_percentage") or 0
            # Also check employer contribution >= 60% of total premium
            total_prem = primary.get("premiums", {}).get("self_only", 0) or 0
            er_contrib = primary.get("employer_contribution", {}).get("self_only", 0) or 0
            er_pct = (er_contrib / total_prem * 100) if total_prem > 0 else 0
            mv_pass = mv_pct >= 60 and er_pct >= 60

            if not is_mec:
                offer_code = "1H"
            elif is_mec and not mv_pass:
                offer_code = "1F"
            else:
                # Determine broadest coverage tier offered
                # Check enrollment tier, plan coverage levels, or employee data
                tier = (enrollment or {}).get("coverage_tier", "") or emp.get("coverage_tier", "")

                # Check plan's coverage_levels to determine broadest offer
                coverage_levels = primary.get("coverage_levels", {})
                has_family = bool(coverage_levels.get("family"))
                has_spouse = bool(coverage_levels.get("employee_spouse"))
                has_deps = bool(coverage_levels.get("employee_children") or coverage_levels.get("employee_dependents"))

                if has_family or tier in ("family", "employee_spouse_dependents"):
                    offer_code = "1E"
                elif has_spouse or tier == "employee_spouse":
                    offer_code = "1D"
                elif has_deps or tier in ("employee_children", "employee_dependents"):
                    offer_code = "1C"
                else:
                    offer_code = "1B"

                # Check for 1A qualifying offer (affordable + MV + offered to spouse/dependents)
                if offer_code == "1E":
                    ee_cost = primary.get("employee_cost", {}).get("self_only", 0)
                    if ee_cost <= 129.89:  # 2026 FPL safe harbor
                        offer_code = "1A"

    await db.eligibility_results.update_one(
        {"employer_id": employer_id, "employee_id": employee_id},
        {"$set": {"offer_code": offer_code}},
    )
    # Also update the employee profile's offer_code
    await db.employee_profiles.update_one(
        {"id": employee_id},
        {"$set": {"offer_code": offer_code}},
    )
    return offer_code



# --- Pydantic Models ---

class PlanLibraryCreate(BaseModel):
    employer_id: str
    carrier_name: str
    plan_name: str
    plan_type: str  # PPO, HMO, HDHP, EPO, POS
    category: str = "medical"  # medical, dental, vision
    premiums_self_only: float = 0
    premiums_employee_spouse: float = 0
    premiums_employee_children: float = 0
    premiums_family: float = 0
    employer_contribution_self_only: float = 0
    employer_contribution_employee_spouse: float = 0
    employer_contribution_employee_children: float = 0
    employer_contribution_family: float = 0
    individual_deductible: float = 0
    family_deductible: float = 0
    coinsurance_rate: float = 0
    oop_max_individual: float = 0
    oop_max_family: float = 0
    copay_primary: float = 0
    copay_specialist: float = 0
    copay_er: float = 0
    copay_generic_rx: float = 0
    copay_brand_rx: float = 0
    mv_percentage: Optional[float] = None
    mv_certified: bool = False
    mec_qualified: bool = True
    plan_year_start: str = ""
    plan_year_end: str = ""
    sbc_url: str = ""


class EnrollmentChoice(BaseModel):
    plan_id: str
    coverage_tier: str  # self_only, employee_spouse, employee_children, family
    add_on_plan_ids: List[str] = []


class DeclineChoice(BaseModel):
    reason: str  # too_expensive, other_coverage, spouse_coverage, medicaid, other
    reason_detail: str = ""


# --- IRS Offer Codes (Line 14 - 1095-C) ---
# These map coverage TIER to the offer code.
# The code reflects what was OFFERED, not what the employee chose.
# 1A = Qualifying Offer (MEC+MV, affordable ≤ FPL, offered to spouse+dependents)
# 1B = MEC+MV offered to employee only
# 1C = MEC+MV offered to employee + dependents (not spouse)
# 1D = MEC+MV offered to employee + spouse (not dependents)
# 1E = MEC+MV offered to employee + spouse + dependents
# 1F = MEC offered but NOT meeting MV (< 60%)
# 1H = No offer of coverage
OFFER_CODE_MAP = {
    "self_only": "1B",
    "individual": "1B",
    "employee_only": "1B",
    "employee_spouse": "1D",
    "employee_children": "1C",
    "employee_dependents": "1C",
    "family": "1E",
    "employee_spouse_dependents": "1E",
}

DECLINE_REASONS = [
    {"value": "too_expensive", "label": "Too expensive"},
    {"value": "other_coverage", "label": "Have other coverage"},
    {"value": "spouse_coverage", "label": "Covered by spouse's plan"},
    {"value": "medicaid", "label": "Enrolled in Medicaid/Medicare"},
    {"value": "marketplace", "label": "Marketplace plan"},
    {"value": "other", "label": "Other reason"},
]


def generate_employer_code():
    """Generate a 6-char alphanumeric employer access code."""
    chars = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    return "".join(random.choices(chars, k=6))


def register_enrollment_routes(router, db, get_current_user):
    """Register all enrollment workflow routes."""

    # =============================================
    # STEP 1: PLAN LIBRARY SETUP (HR ADMIN)
    # =============================================

    @router.post("/enrollment/plans")
    async def create_plan_library_entry(data: PlanLibraryCreate, user=Depends(get_current_user)):
        """HR creates a plan in the library."""
        plan_id = str(uuid.uuid4())
        ee_cost_self = data.premiums_self_only - data.employer_contribution_self_only
        plan_doc = {
            "id": plan_id,
            "employer_id": data.employer_id,
            "carrier_name": data.carrier_name,
            "plan_name": data.plan_name,
            "plan_type": data.plan_type,
            "category": data.category,
            "premiums": {
                "self_only": data.premiums_self_only,
                "employee_spouse": data.premiums_employee_spouse,
                "employee_children": data.premiums_employee_children,
                "family": data.premiums_family,
            },
            "employer_contribution": {
                "self_only": data.employer_contribution_self_only,
                "employee_spouse": data.employer_contribution_employee_spouse,
                "employee_children": data.employer_contribution_employee_children,
                "family": data.employer_contribution_family,
            },
            "employee_cost": {
                "self_only": max(0, ee_cost_self),
                "employee_spouse": max(0, data.premiums_employee_spouse - data.employer_contribution_employee_spouse),
                "employee_children": max(0, data.premiums_employee_children - data.employer_contribution_employee_children),
                "family": max(0, data.premiums_family - data.employer_contribution_family),
            },
            "individual_deductible": data.individual_deductible,
            "family_deductible": data.family_deductible,
            "coinsurance_rate": data.coinsurance_rate,
            "oop_max_individual": data.oop_max_individual,
            "oop_max_family": data.oop_max_family,
            "copay_primary": data.copay_primary,
            "copay_specialist": data.copay_specialist,
            "copay_er": data.copay_er,
            "copay_generic_rx": data.copay_generic_rx,
            "copay_brand_rx": data.copay_brand_rx,
            "mv_percentage": data.mv_percentage,
            "mv_certified": data.mv_certified,
            "mec_qualified": data.mec_qualified,
            "plan_year_start": data.plan_year_start,
            "plan_year_end": data.plan_year_end,
            "sbc_url": data.sbc_url,
            "status": "active",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.plan_library.insert_one(plan_doc)
        plan_doc.pop("_id", None)
        return plan_doc

    @router.get("/enrollment/plans/{employer_id}")
    async def get_plan_library(employer_id: str, user=Depends(get_current_user)):
        """Get all plans in library for employer."""
        plans = await db.plan_library.find(
            {"employer_id": employer_id}, {"_id": 0}
        ).sort("category", 1).to_list(100)
        return plans

    @router.put("/enrollment/plans/{plan_id}")
    async def update_plan_library_entry(plan_id: str, data: dict, user=Depends(get_current_user)):
        """Update a plan in the library."""
        plan = await db.plan_library.find_one({"id": plan_id}, {"_id": 0})
        if not plan:
            raise HTTPException(status_code=404, detail="Plan not found")

        allowed = [
            "carrier_name", "plan_name", "plan_type", "category",
            "premiums_self_only", "premiums_employee_spouse", "premiums_employee_children", "premiums_family",
            "employer_contribution_self_only", "employer_contribution_employee_spouse",
            "employer_contribution_employee_children", "employer_contribution_family",
            "individual_deductible", "family_deductible", "coinsurance_rate",
            "oop_max_individual", "oop_max_family", "copay_primary", "copay_specialist",
            "copay_er", "copay_generic_rx", "copay_brand_rx",
            "mv_percentage", "mv_certified", "mec_qualified",
            "plan_year_start", "plan_year_end", "sbc_url", "status",
        ]
        update = {k: v for k, v in data.items() if k in allowed}

        # Recalculate premiums structure if tier premiums provided
        p_so = data.get("premiums_self_only", plan.get("premiums", {}).get("self_only", 0))
        p_es = data.get("premiums_employee_spouse", plan.get("premiums", {}).get("employee_spouse", 0))
        p_ec = data.get("premiums_employee_children", plan.get("premiums", {}).get("employee_children", 0))
        p_f = data.get("premiums_family", plan.get("premiums", {}).get("family", 0))
        ec_so = data.get("employer_contribution_self_only", plan.get("employer_contribution", {}).get("self_only", 0))
        ec_es = data.get("employer_contribution_employee_spouse", plan.get("employer_contribution", {}).get("employee_spouse", 0))
        ec_ec = data.get("employer_contribution_employee_children", plan.get("employer_contribution", {}).get("employee_children", 0))
        ec_f = data.get("employer_contribution_family", plan.get("employer_contribution", {}).get("family", 0))

        update["premiums"] = {"self_only": p_so, "employee_spouse": p_es, "employee_children": p_ec, "family": p_f}
        update["employer_contribution"] = {"self_only": ec_so, "employee_spouse": ec_es, "employee_children": ec_ec, "family": ec_f}
        update["employee_cost"] = {
            "self_only": max(0, p_so - ec_so),
            "employee_spouse": max(0, p_es - ec_es),
            "employee_children": max(0, p_ec - ec_ec),
            "family": max(0, p_f - ec_f),
        }
        # Remove flat premium keys from update
        for k in ["premiums_self_only", "premiums_employee_spouse", "premiums_employee_children", "premiums_family",
                   "employer_contribution_self_only", "employer_contribution_employee_spouse",
                   "employer_contribution_employee_children", "employer_contribution_family"]:
            update.pop(k, None)

        update["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.plan_library.update_one({"id": plan_id}, {"$set": update})
        result = await db.plan_library.find_one({"id": plan_id}, {"_id": 0})
        return result

    @router.delete("/enrollment/plans/{plan_id}")
    async def delete_plan_library_entry(plan_id: str, user=Depends(get_current_user)):
        result = await db.plan_library.delete_one({"id": plan_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Plan not found")
        return {"message": "Plan deleted"}

    @router.post("/enrollment/plans/upload/{employer_id}")
    async def upload_plans_csv(employer_id: str, file: UploadFile = File(...), user=Depends(get_current_user)):
        """Upload plans from CSV file."""
        content = await file.read()
        text = content.decode("utf-8")
        reader = csv.DictReader(io.StringIO(text))

        created = []
        errors = []
        for i, row in enumerate(reader):
            try:
                plan_id = str(uuid.uuid4())
                p_so = float(row.get("premiums_self_only", 0) or 0)
                ec_so = float(row.get("employer_contribution_self_only", 0) or 0)
                p_es = float(row.get("premiums_employee_spouse", 0) or 0)
                ec_es = float(row.get("employer_contribution_employee_spouse", 0) or 0)
                p_ec = float(row.get("premiums_employee_children", 0) or 0)
                ec_ec = float(row.get("employer_contribution_employee_children", 0) or 0)
                p_f = float(row.get("premiums_family", 0) or 0)
                ec_f = float(row.get("employer_contribution_family", 0) or 0)

                plan_doc = {
                    "id": plan_id,
                    "employer_id": employer_id,
                    "carrier_name": row.get("carrier_name", ""),
                    "plan_name": row.get("plan_name", f"Plan {i+1}"),
                    "plan_type": row.get("plan_type", "PPO"),
                    "category": row.get("category", "medical"),
                    "premiums": {"self_only": p_so, "employee_spouse": p_es, "employee_children": p_ec, "family": p_f},
                    "employer_contribution": {"self_only": ec_so, "employee_spouse": ec_es, "employee_children": ec_ec, "family": ec_f},
                    "employee_cost": {
                        "self_only": max(0, p_so - ec_so), "employee_spouse": max(0, p_es - ec_es),
                        "employee_children": max(0, p_ec - ec_ec), "family": max(0, p_f - ec_f),
                    },
                    "individual_deductible": float(row.get("individual_deductible", 0) or 0),
                    "family_deductible": float(row.get("family_deductible", 0) or 0),
                    "coinsurance_rate": float(row.get("coinsurance_rate", 0) or 0),
                    "oop_max_individual": float(row.get("oop_max_individual", 0) or 0),
                    "oop_max_family": float(row.get("oop_max_family", 0) or 0),
                    "copay_primary": float(row.get("copay_primary", 0) or 0),
                    "copay_specialist": float(row.get("copay_specialist", 0) or 0),
                    "copay_er": float(row.get("copay_er", 0) or 0),
                    "copay_generic_rx": float(row.get("copay_generic_rx", 0) or 0),
                    "copay_brand_rx": float(row.get("copay_brand_rx", 0) or 0),
                    "mv_percentage": float(row.get("mv_percentage", 0) or 0) or None,
                    "mv_certified": str(row.get("mv_certified", "")).lower() in ("true", "1", "yes"),
                    "mec_qualified": str(row.get("mec_qualified", "true")).lower() in ("true", "1", "yes"),
                    "plan_year_start": row.get("plan_year_start", ""),
                    "plan_year_end": row.get("plan_year_end", ""),
                    "sbc_url": row.get("sbc_url", ""),
                    "status": "active",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                }
                await db.plan_library.insert_one(plan_doc)
                plan_doc.pop("_id", None)
                created.append(plan_doc)
            except Exception as e:
                errors.append({"row": i + 1, "error": str(e)})

        return {"created": len(created), "errors": errors, "plans": created}

    # =============================================
    # PLAN COMPLIANCE CHECK & EMPLOYEE ASSIGNMENT
    # =============================================

    @router.post("/enrollment/plans/{plan_id}/compliance-check")
    async def check_plan_compliance(plan_id: str, user=Depends(get_current_user)):
        """Run MEC, MV, and Affordability compliance checks on a plan."""
        plan = await db.plan_library.find_one({"id": plan_id}, {"_id": 0})
        if not plan:
            raise HTTPException(status_code=404, detail="Plan not found")

        employer_id = plan["employer_id"]

        # MEC Check
        mec_checks = {
            "is_group_health_plan": True,
            "covers_essential_benefits": plan.get("mec_qualified", False),
            "covers_preventive_care": True,
            "no_annual_limits": plan.get("oop_max_individual", 0) > 0,
            "no_lifetime_limits": True,
            "covers_dependents_to_26": True,
            "no_preexisting_exclusions": True,
        }
        mec_pass = all(mec_checks.values())

        # MV Check — Dynamic HHS MV Calculator
        # Uses standard population costs and plan cost-sharing to calculate actuarial value.
        # Also checks employer contribution >= 60% of total premium.
        hhs_result = hhs_mv_calculate(plan)
        mv_percentage = hhs_result["mv_percentage"]
        er_contrib_pct = hhs_result["premium_analysis"]["employer_pct"]

        mv_actuarial_pass = hhs_result["meets_minimum"]  # mv_percentage >= 60
        mv_contribution_pass = hhs_result["premium_analysis"]["employer_contribution_pass"]
        mv_pass = hhs_result["overall_mv_pass"]

        # Check for actual actuarial certification via quote_requests
        actuary_cert = None
        cert_quote = await db.quote_requests.find_one(
            {"plan_id": plan_id, "status": {"$in": ["delivered", "validated", "completed"]}},
            {"_id": 0, "certification": 1, "actuary_name": 1, "status": 1}
        )
        if cert_quote and cert_quote.get("certification"):
            cert = cert_quote["certification"]
            actuary_cert = {
                "certified": True,
                "actuary_name": cert_quote.get("actuary_name", ""),
                "mv_percentage": cert.get("mv_percentage", 0),
                "notes": cert.get("certification_notes", ""),
                "delivered_at": cert.get("delivered_at", ""),
                "pass": cert.get("mv_percentage", 0) >= 60,
            }

        # Also check if there's a failed/rejected certification
        if not actuary_cert:
            failed_quote = await db.quote_requests.find_one(
                {"plan_id": plan_id, "status": "resubmit_needed"},
                {"_id": 0, "certification": 1, "actuary_name": 1, "validation": 1}
            )
            if failed_quote and failed_quote.get("certification"):
                cert = failed_quote["certification"]
                actuary_cert = {
                    "certified": False,
                    "actuary_name": failed_quote.get("actuary_name", ""),
                    "mv_percentage": cert.get("mv_percentage", 0),
                    "notes": cert.get("certification_notes", ""),
                    "delivered_at": cert.get("delivered_at", ""),
                    "pass": False,
                    "rejection_reason": (failed_quote.get("validation") or {}).get("rejection_reason", ""),
                }

        # Check if there's any active/pending quote for this plan
        active_quote = await db.quote_requests.find_one(
            {"plan_id": plan_id, "status": {"$in": ["pending", "accepted", "paid", "delivered"]}},
            {"_id": 0, "id": 1, "status": 1}
        )

        mv_detail = {
            "mv_percentage": mv_percentage,
            "threshold": 60,
            "employer_contribution_pct": er_contrib_pct,
            "employer_contribution_pass": mv_contribution_pass,
            "actuarial_pass": mv_actuarial_pass,
            "total_allowed_cost": hhs_result["total_allowed_cost"],
            "plan_pays": hhs_result["plan_pays"],
            "member_pays": hhs_result["member_pays"],
            "oop_max_applied": hhs_result["oop_max_applied"],
            "category_breakdown": hhs_result["category_breakdown"],
            "premium_analysis": hhs_result["premium_analysis"],
            "calculation_notes": hhs_result["calculation_notes"],
            "plan_parameters": {
                "deductible": hhs_result["deductible_used"],
                "coinsurance": hhs_result["coinsurance_used"],
                "oop_max": hhs_result["oop_max_used"],
            },
            "certified_by_actuary": actuary_cert,
            "has_active_quote": bool(active_quote),
            "active_quote_status": active_quote.get("status") if active_quote else None,
            "certification_source": plan.get("certification_source"),
            "certification_status": plan.get("certification_status"),
            "method": "HHS MV Calculator",
        }

        # Affordability Check across full-time employees & FTEs only (ACA requirement)
        employees = await db.employee_profiles.find(
            {"employer_id": employer_id}, {"_id": 0}
        ).to_list(500)
        if not employees:
            employees = await db.payroll_employees.find(
                {"employer_id": employer_id}, {"_id": 0}
            ).to_list(500)

        # Filter: only full-time (30+ hrs/wk) or FTE (130+ monthly hrs)
        ft_employees = []
        for emp in employees:
            weekly_hrs = emp.get("weekly_hours", 0)
            monthly_hrs = emp.get("monthly_hours", weekly_hrs * 4.33)
            emp_type = emp.get("employment_type", "")
            if weekly_hrs >= 30 or monthly_hrs >= 130 or emp_type == "full_time":
                ft_employees.append(emp)

        ee_cost = plan.get("employee_cost", {}).get("self_only", 0)
        annual_ee_cost = ee_cost * 12
        affordable_count = 0
        unaffordable_count = 0
        total_checked = 0
        employee_affordability = []
        for emp in ft_employees:
            salary = emp.get("annual_salary", 0) or emp.get("w2_wages", 0)
            if salary > 0:
                total_checked += 1
                threshold = salary * 0.0996
                is_affordable = annual_ee_cost <= threshold
                if is_affordable:
                    affordable_count += 1
                else:
                    unaffordable_count += 1
                employee_affordability.append({
                    "employee_id": emp.get("id"),
                    "name": emp.get("name", "Unknown"),
                    "annual_salary": round(salary, 2),
                    "monthly_threshold": round(threshold / 12, 2),
                    "employee_monthly_cost": ee_cost,
                    "pct_of_income": round((annual_ee_cost / salary) * 100, 2),
                    "affordable": is_affordable,
                })

        # Sort: unaffordable first, then by pct_of_income descending
        employee_affordability.sort(key=lambda x: (x["affordable"], -x["pct_of_income"]))

        affordability = {
            "employee_monthly_cost": ee_cost,
            "annual_cost": annual_ee_cost,
            "threshold_rate": 9.96,
            "method": "W-2 Safe Harbor",
            "total_ft_employees": len(ft_employees),
            "total_employees_checked": total_checked,
            "affordable_for": affordable_count,
            "unaffordable_for": unaffordable_count,
            "pass_rate": round((affordable_count / total_checked * 100), 1) if total_checked > 0 else 0,
            "employees": employee_affordability,
        }

        # Update stored mv_percentage with HHS-calculated value
        # mv_certified reflects OVERALL pass: both actuarial value >= 60% AND employer contribution >= 60%
        await db.plan_library.update_one(
            {"id": plan_id},
            {"$set": {"mv_percentage": mv_percentage, "mv_certified": mv_pass, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )

        return {
            "plan_id": plan_id,
            "plan_name": plan["plan_name"],
            "mec": {"pass": mec_pass, "checks": mec_checks},
            "mv": {"pass": mv_pass, **mv_detail},
            "affordability": affordability,
            "overall_compliant": mec_pass and mv_pass and (affordability["pass_rate"] >= 95 if total_checked > 0 else True),
        }

    @router.get("/enrollment/plans/{plan_id}/assigned-employees")
    async def get_assigned_employees(plan_id: str, user=Depends(get_current_user)):
        """Get employees assigned to a specific plan."""
        assignments = await db.plan_assignments.find(
            {"plan_id": plan_id}, {"_id": 0}
        ).to_list(500)
        return assignments

    @router.post("/enrollment/plans/{plan_id}/assign-employees")
    async def assign_employees_to_plan(plan_id: str, data: dict, user=Depends(get_current_user)):
        """Assign selected employees to a plan."""
        plan = await db.plan_library.find_one({"id": plan_id}, {"_id": 0})
        if not plan:
            raise HTTPException(status_code=404, detail="Plan not found")

        employee_ids = data.get("employee_ids", [])
        if not employee_ids:
            raise HTTPException(status_code=400, detail="No employees selected")

        employer_id = plan["employer_id"]

        # Affordability check for each employee before assigning
        ee_cost = plan.get("employee_cost", {}).get("self_only", 0) or 0
        annual_ee_cost = ee_cost * 12
        unaffordable = []
        affordable_ids = []

        if plan.get("category") == "medical" and ee_cost > 0:
            for emp_id in employee_ids:
                emp = await db.employee_profiles.find_one({"id": emp_id}, {"_id": 0})
                if not emp:
                    emp = await db.payroll_employees.find_one({"id": emp_id}, {"_id": 0})
                if not emp:
                    affordable_ids.append(emp_id)
                    continue
                salary = emp.get("annual_salary", 0) or emp.get("w2_wages", 0)
                if salary > 0:
                    threshold = salary * 0.0996
                    pct = round((annual_ee_cost / salary) * 100, 2)
                    if annual_ee_cost > threshold:
                        unaffordable.append({
                            "employee_id": emp_id,
                            "name": emp.get("name", "Unknown"),
                            "annual_salary": salary,
                            "employee_monthly_cost": ee_cost,
                            "max_affordable_monthly": round(threshold / 12, 2),
                            "pct_of_income": pct,
                            "threshold": 9.96,
                        })
                    else:
                        affordable_ids.append(emp_id)
                else:
                    affordable_ids.append(emp_id)
        else:
            affordable_ids = list(employee_ids)

        # If ALL requested employees are unaffordable, block the assignment entirely
        if unaffordable and not affordable_ids:
            names = ", ".join(u["name"] for u in unaffordable[:3])
            extra = f" and {len(unaffordable) - 3} more" if len(unaffordable) > 3 else ""
            raise HTTPException(status_code=422, detail={
                "message": f"Plan is unaffordable for all {len(unaffordable)} selected employee(s): {names}{extra}. The employee cost of ${ee_cost}/mo exceeds the ACA 9.96% affordability threshold based on their salary.",
                "unaffordable_employees": unaffordable,
                "plan_cost": ee_cost,
            })

        # Proceed with affordable employees only
        assigned_ids = affordable_ids if affordable_ids else employee_ids

        assigned = 0
        skipped_unaffordable = len(unaffordable)
        for emp_id in assigned_ids:
            # Get employee info
            emp = await db.employee_profiles.find_one({"id": emp_id}, {"_id": 0})
            if not emp:
                emp = await db.payroll_employees.find_one({"id": emp_id}, {"_id": 0})
            if not emp:
                continue

            assignment = {
                "id": str(uuid.uuid4()),
                "employer_id": employer_id,
                "plan_id": plan_id,
                "plan_name": plan["plan_name"],
                "plan_category": plan.get("category", "medical"),
                "employee_id": emp_id,
                "employee_name": emp.get("name", ""),
                "assigned_at": datetime.now(timezone.utc).isoformat(),
                "assigned_by": user["id"],
            }
            await db.plan_assignments.update_one(
                {"plan_id": plan_id, "employee_id": emp_id},
                {"$set": assignment},
                upsert=True,
            )
            assigned += 1

            # Recalculate offer code for this employee
            await _recalculate_offer_code(db, employer_id, emp_id, emp)

            # Notify employee about plan assignment
            emp_user = await db.users.find_one({"linked_employee_id": emp_id}, {"_id": 0, "id": 1})
            if emp_user:
                await _create_notification(db, emp_user["id"],
                    "Plan Assigned to You",
                    f"You have been assigned to {plan['plan_name']}. Please review and accept or decline during the enrollment window.",
                    category="assignment", link="/employee-portal")

        return {"assigned": assigned, "skipped_unaffordable": skipped_unaffordable, "plan_name": plan["plan_name"]}

    @router.post("/enrollment/plans/{plan_id}/unassign-employees")
    async def unassign_employees_from_plan(plan_id: str, data: dict, user=Depends(get_current_user)):
        """Remove employee assignments from a plan."""
        employee_ids = data.get("employee_ids", [])
        if not employee_ids:
            raise HTTPException(status_code=400, detail="No employees selected")

        plan = await db.plan_library.find_one({"id": plan_id}, {"_id": 0})
        employer_id = plan["employer_id"] if plan else None

        result = await db.plan_assignments.delete_many(
            {"plan_id": plan_id, "employee_id": {"$in": employee_ids}}
        )

        # Recalculate offer codes for unassigned employees
        if employer_id:
            for emp_id in employee_ids:
                emp = await db.employee_profiles.find_one({"id": emp_id}, {"_id": 0})
                if emp:
                    await _recalculate_offer_code(db, employer_id, emp_id, emp)

        return {"unassigned": result.deleted_count}

    @router.get("/enrollment/assignments/{employer_id}")
    async def get_all_assignments(employer_id: str, user=Depends(get_current_user)):
        """Get all plan assignments for an employer."""
        assignments = await db.plan_assignments.find(
            {"employer_id": employer_id}, {"_id": 0}
        ).to_list(1000)
        return assignments

    @router.get("/enrollment/employees-list/{employer_id}")
    async def get_employees_for_assignment(employer_id: str, user=Depends(get_current_user)):
        """Get employee list for plan assignment UI."""
        employees = await db.employee_profiles.find(
            {"employer_id": employer_id},
            {"_id": 0, "id": 1, "name": 1, "email": 1, "department": 1, "employment_type": 1,
             "weekly_hours": 1, "annual_salary": 1}
        ).sort("name", 1).to_list(500)
        if not employees:
            employees = await db.payroll_employees.find(
                {"employer_id": employer_id},
                {"_id": 0, "id": 1, "name": 1, "email": 1, "department": 1, "employment_type": 1,
                 "weekly_hours": 1, "annual_salary": 1}
            ).sort("name", 1).to_list(500)
        return employees

    # =============================================
    # STEP 2: AUTO-ELIGIBILITY ENGINE
    # =============================================

    @router.post("/enrollment/eligibility/run/{employer_id}")
    async def run_eligibility_engine(employer_id: str, user=Depends(get_current_user)):
        """Run eligibility engine for all employees of an employer."""
        # Get employees
        employees = await db.employee_profiles.find(
            {"employer_id": employer_id}, {"_id": 0}
        ).to_list(500)
        if not employees:
            employees = await db.payroll_employees.find(
                {"employer_id": employer_id}, {"_id": 0}
            ).to_list(500)
        if not employees:
            raise HTTPException(status_code=400, detail="No employees found")

        # Get plan library
        plans = await db.plan_library.find(
            {"employer_id": employer_id, "status": "active"}, {"_id": 0}
        ).to_list(100)

        medical_plans = [p for p in plans if p.get("category") == "medical"]
        addon_plans = [p for p in plans if p.get("category") in ("dental", "vision")]

        results = []
        for emp in employees:
            weekly_hours = emp.get("weekly_hours", 0)
            monthly_hours = emp.get("monthly_hours", weekly_hours * 4.33)
            is_ft = monthly_hours >= 130 or weekly_hours >= 30
            annual_salary = emp.get("annual_salary", 0)

            # Determine eligible plans (MEC-qualified medical plans)
            eligible_plan_ids = []
            for p in medical_plans:
                if p.get("mec_qualified", True):
                    eligible_plan_ids.append(p["id"])

            # W-2 safe harbor affordability check
            affordable = False
            contribution = 0
            if eligible_plan_ids and annual_salary > 0:
                # Use lowest-cost self-only plan
                lowest_ee_cost = min(
                    (p["employee_cost"]["self_only"] for p in medical_plans if p["id"] in eligible_plan_ids),
                    default=0,
                )
                contribution = lowest_ee_cost
                annual_ee_cost = lowest_ee_cost * 12
                # 2026 affordability threshold: 9.96%
                threshold = annual_salary * 0.0996
                affordable = annual_ee_cost <= threshold

            # Offer code: based on plan assignments OR enrollments (accepted/declined)
            employee_assignments = await db.plan_assignments.find(
                {"employer_id": employer_id, "employee_id": emp["id"]}, {"_id": 0}
            ).to_list(20)
            assigned_plan_ids = [a["plan_id"] for a in employee_assignments]

            # Also check enrollments — if employee enrolled or declined, they were offered
            employee_enrollment = await db.enrollments.find_one(
                {"employer_id": employer_id, "employee_id": emp["id"]}, {"_id": 0}
            )
            if employee_enrollment and employee_enrollment.get("plan_id") and employee_enrollment["plan_id"] not in assigned_plan_ids:
                assigned_plan_ids.append(employee_enrollment["plan_id"])

            if not is_ft:
                offer_code = "1H"  # Not full-time
            elif not assigned_plan_ids:
                offer_code = "1H"  # No plan assigned or enrollment record
            else:
                # Get assigned plan details from plan_library
                assigned_plans_data = []
                for pid in assigned_plan_ids:
                    p = await db.plan_library.find_one({"id": pid}, {"_id": 0})
                    if p:
                        assigned_plans_data.append(p)
                medical_assigned = [p for p in assigned_plans_data if p.get("category") == "medical"]
                if not medical_assigned:
                    offer_code = "1H"  # No medical plan available
                else:
                    primary = medical_assigned[0]
                    is_mec = primary.get("mec_qualified", False)
                    mv_pct = primary.get("mv_percentage") or 0
                    total_prem = primary.get("premiums", {}).get("self_only", 0) or 0
                    er_contrib = primary.get("employer_contribution", {}).get("self_only", 0) or 0
                    er_pct = (er_contrib / total_prem * 100) if total_prem > 0 else 0
                    mv_pass = mv_pct >= 60 and er_pct >= 60
                    tier = emp.get("coverage_tier", "individual")
                    if not is_mec:
                        offer_code = "1H"  # Non-MEC
                    elif is_mec and not mv_pass:
                        offer_code = "1F"  # MEC but MV < 60%
                    elif tier in ("family", "employee_spouse_dependents"):
                        offer_code = "1E"  # MEC + MV to employee, spouse & dependents
                    elif tier == "employee_spouse":
                        offer_code = "1D"  # MEC + MV to employee + spouse
                    elif tier in ("employee_children", "employee_dependents"):
                        offer_code = "1C"  # MEC + MV to employee + dependents
                    else:
                        offer_code = "1B"  # MEC + MV to employee only

            eligibility = {
                "id": str(uuid.uuid4()),
                "employer_id": employer_id,
                "employee_id": emp["id"],
                "employee_name": emp.get("name", ""),
                "is_full_time": is_ft,
                "weekly_hours": weekly_hours,
                "monthly_hours": round(monthly_hours, 1),
                "annual_salary": annual_salary,
                "eligible": is_ft and len(eligible_plan_ids) > 0,
                "eligible_plan_ids": eligible_plan_ids if is_ft else [],
                "addon_plan_ids": [p["id"] for p in addon_plans] if is_ft else [],
                "offer_code": offer_code if is_ft else "1H",  # 1H = no offer
                "safe_harbor_method": "W-2",
                "affordable": affordable,
                "lowest_cost_contribution": contribution,
                "calculated_at": datetime.now(timezone.utc).isoformat(),
            }
            results.append(eligibility)

        # Upsert eligibility results
        for r in results:
            await db.eligibility_results.update_one(
                {"employer_id": employer_id, "employee_id": r["employee_id"]},
                {"$set": r},
                upsert=True,
            )

        eligible_count = sum(1 for r in results if r["eligible"])
        return {
            "total_employees": len(results),
            "eligible": eligible_count,
            "ineligible": len(results) - eligible_count,
            "plans_available": len(medical_plans),
            "addon_plans": len(addon_plans),
            "results": results,
        }

    @router.get("/enrollment/eligibility/{employer_id}")
    async def get_eligibility_results(employer_id: str, user=Depends(get_current_user)):
        """Get cached eligibility results."""
        results = await db.eligibility_results.find(
            {"employer_id": employer_id}, {"_id": 0}
        ).to_list(500)
        eligible_count = sum(1 for r in results if r.get("eligible"))
        return {
            "total": len(results),
            "eligible": eligible_count,
            "ineligible": len(results) - eligible_count,
            "results": results,
        }

    # =============================================
    # EMPLOYER CODE FOR EMPLOYEE SELF-REGISTRATION
    # =============================================

    @router.get("/enrollment/employer-code/{employer_id}")
    async def get_employer_code(employer_id: str, user=Depends(get_current_user)):
        """Get or generate employer access code for employee registration."""
        employer = await db.employers.find_one({"id": employer_id}, {"_id": 0})
        if not employer:
            raise HTTPException(status_code=404, detail="Employer not found")

        code = employer.get("access_code")
        if not code:
            code = generate_employer_code()
            await db.employers.update_one({"id": employer_id}, {"$set": {"access_code": code}})

        return {"access_code": code, "employer_name": employer["name"]}

    @router.post("/enrollment/employee/register")
    async def register_employee_user(data: dict):
        """Employee self-registers with employer access code."""
        email = data.get("email", "")
        password = data.get("password", "")
        name = data.get("name", "")
        employer_code = data.get("employer_code", "")

        if not all([email, password, name, employer_code]):
            raise HTTPException(status_code=400, detail="All fields required")

        # Find employer by code
        employer = await db.employers.find_one({"access_code": employer_code}, {"_id": 0})
        if not employer:
            raise HTTPException(status_code=400, detail="Invalid employer code")

        # Check if email already registered
        existing = await db.users.find_one({"email": email}, {"_id": 0})
        if existing:
            raise HTTPException(status_code=400, detail="Email already registered")

        import bcrypt
        user_id = str(uuid.uuid4())
        hashed = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        user_doc = {
            "id": user_id,
            "email": email,
            "password": hashed,
            "name": name,
            "role": "employee",
            "employer_id": employer["id"],
            "employer_name": employer["name"],
            "company_name": employer["name"],
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.users.insert_one(user_doc)

        # Try to link to existing employee record
        emp_match = await db.employee_profiles.find_one(
            {"employer_id": employer["id"], "email": email}, {"_id": 0}
        )
        if not emp_match:
            emp_match = await db.payroll_employees.find_one(
                {"employer_id": employer["id"], "name": {"$regex": name, "$options": "i"}}, {"_id": 0}
            )

        if emp_match:
            await db.users.update_one(
                {"id": user_id},
                {"$set": {"linked_employee_id": emp_match["id"]}}
            )

        import jwt as pyjwt
        import os
        JWT_SECRET = os.environ.get("JWT_SECRET", "complicore-secret-key-2025")
        token = pyjwt.encode(
            {"user_id": user_id, "role": "employee", "exp": datetime.now(timezone.utc).timestamp() + 86400 * 7},
            JWT_SECRET, algorithm="HS256",
        )

        return {
            "token": token,
            "user": {
                "id": user_id, "email": email, "name": name,
                "role": "employee", "employer_id": employer["id"],
                "employer_name": employer["name"],
                "company_name": employer["name"],
                "linked_employee_id": emp_match["id"] if emp_match else None,
                "created_at": user_doc["created_at"],
            },
        }

    # =============================================
    # STEP 3: EMPLOYEE SELF-SERVICE PORTAL
    # =============================================

    @router.get("/enrollment/employee/my-plans")
    async def get_employee_eligible_plans(user=Depends(get_current_user)):
        """Employee views their eligible plans."""
        emp_id = user.get("linked_employee_id")
        employer_id = user.get("employer_id")

        if not employer_id:
            raise HTTPException(status_code=400, detail="No employer linked")

        # Get eligibility
        eligibility = None
        if emp_id:
            eligibility = await db.eligibility_results.find_one(
                {"employee_id": emp_id}, {"_id": 0}
            )

        # Get all plans for this employer
        plans = await db.plan_library.find(
            {"employer_id": employer_id, "status": "active"}, {"_id": 0}
        ).to_list(100)

        # Get existing enrollment
        enrollment = await db.enrollments.find_one(
            {"employer_id": employer_id, "user_id": user["id"]}, {"_id": 0}
        )

        # Check if this employee has specific plan assignments
        emp_id_for_assignment = user.get("linked_employee_id") or user.get("id")
        assignments = await db.plan_assignments.find(
            {"employer_id": employer_id, "employee_id": emp_id_for_assignment}, {"_id": 0}
        ).to_list(100)

        if assignments:
            # Use assigned plans only
            assigned_plan_ids = [a["plan_id"] for a in assignments]
            medical_plans = [p for p in plans if p["id"] in assigned_plan_ids and p.get("category") == "medical"]
            addon_plans = [p for p in plans if p["id"] in assigned_plan_ids and p.get("category") in ("dental", "vision")]
        else:
            # Fallback to eligibility-based
            eligible_plan_ids = eligibility.get("eligible_plan_ids", []) if eligibility else [p["id"] for p in plans if p.get("category") == "medical"]
            addon_ids = eligibility.get("addon_plan_ids", []) if eligibility else [p["id"] for p in plans if p.get("category") in ("dental", "vision")]
            medical_plans = [p for p in plans if p["id"] in eligible_plan_ids]
            addon_plans = [p for p in plans if p["id"] in addon_ids]

        # Enrich enrollment with full plan details if enrolled
        enrolled_plan_detail = None
        enrolled_addon_details = []
        if enrollment and enrollment.get("plan_id"):
            enrolled_plan_detail = await db.plan_library.find_one(
                {"id": enrollment["plan_id"]}, {"_id": 0}
            )
        if enrollment and enrollment.get("add_ons"):
            for ao in enrollment["add_ons"]:
                ao_plan = await db.plan_library.find_one({"id": ao["plan_id"]}, {"_id": 0})
                if ao_plan:
                    enrolled_addon_details.append(ao_plan)

        return {
            "eligible": eligibility.get("eligible", True) if eligibility else True,
            "offer_code": eligibility.get("offer_code", "") if eligibility else "",
            "medical_plans": medical_plans,
            "addon_plans": addon_plans,
            "current_enrollment": enrollment,
            "enrolled_plan_detail": enrolled_plan_detail,
            "enrolled_addon_details": enrolled_addon_details,
            "decline_reasons": DECLINE_REASONS,
            "employer_name": user.get("employer_name", ""),
            "employer_id": employer_id,
        }

    @router.get("/enrollment/employee/my-1095c/pdf")
    async def get_employee_1095c_pdf(user=Depends(get_current_user)):
        """Employee downloads their own Form 1095-C PDF."""
        from services.irs_forms import generate_1095c_data, render_1095c_pdf, OFFER_CODES

        employer_id = user.get("employer_id")
        emp_id = user.get("linked_employee_id")

        if not employer_id:
            raise HTTPException(status_code=400, detail="No employer linked")

        employer = await db.employers.find_one({"id": employer_id}, {"_id": 0})
        if not employer:
            raise HTTPException(status_code=404, detail="Employer not found")

        # Find employee profile
        employee = None
        if emp_id:
            employee = await db.employee_profiles.find_one({"id": emp_id}, {"_id": 0})

        if not employee:
            raise HTTPException(status_code=404, detail="Employee profile not found. Contact your HR administrator.")

        # Check eligibility/offer code
        elig = await db.eligibility_results.find_one(
            {"employer_id": employer_id, "employee_id": emp_id}, {"_id": 0}
        )
        offer_code = elig.get("offer_code", "1H") if elig else "1H"

        # Also check enrollment record as proof of offer
        enrollment = await db.enrollments.find_one(
            {"employer_id": employer_id, "user_id": user["id"]}, {"_id": 0}
        )
        if offer_code == "1H" and enrollment and enrollment.get("status") in ("enrolled", "declined"):
            offer_code = enrollment.get("offer_code", "1H")

        if offer_code == "1H":
            raise HTTPException(status_code=400, detail="No 1095-C available: you were not offered coverage.")

        # Get plan info
        plan = {}
        if employee.get("plan_id"):
            plan = await db.plan_library.find_one({"id": employee["plan_id"]}, {"_id": 0}) or {}

        tax_year = datetime.now(timezone.utc).year
        form_data = generate_1095c_data(employee, employer, plan, tax_year)

        # Override Line 14 with offer code
        form_data["part2"]["line14_all_year"] = offer_code
        for m in form_data["part2"]["monthly_data"]:
            m["line14_code"] = offer_code
            m["line14_description"] = OFFER_CODES.get(offer_code, "")

        pdf_bytes = render_1095c_pdf(form_data)

        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=1095-C_{tax_year}.pdf"}
        )


    @router.post("/enrollment/employee/enroll")
    async def enroll_employee(data: EnrollmentChoice, user=Depends(get_current_user)):
        """Employee enrolls in a plan."""
        employer_id = user.get("employer_id")
        if not employer_id:
            raise HTTPException(status_code=400, detail="No employer linked")

        # Check enrollment window
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        active_period = await db.enrollment_periods.find_one(
            {"employer_id": employer_id, "status": "active"}, {"_id": 0}
        )
        if active_period:
            if today < active_period["start_date"] or today > active_period["end_date"]:
                # Check exception
                exc = await db.enrollment_exceptions.find_one(
                    {"employer_id": employer_id, "employee_user_id": user["id"], "status": "approved"}, {"_id": 0}
                )
                if not exc:
                    raise HTTPException(status_code=403, detail="Enrollment window is closed. Request an exception from your employer.")
        # Also block if there's a closed period and no active one (and periods exist)
        if not active_period:
            any_period = await db.enrollment_periods.find_one({"employer_id": employer_id}, {"_id": 0})
            if any_period:
                exc = await db.enrollment_exceptions.find_one(
                    {"employer_id": employer_id, "employee_user_id": user["id"], "status": "approved"}, {"_id": 0}
                )
                if not exc:
                    raise HTTPException(status_code=403, detail="Enrollment window is closed. Request an exception from your employer.")

        plan = await db.plan_library.find_one({"id": data.plan_id}, {"_id": 0})
        if not plan:
            raise HTTPException(status_code=404, detail="Plan not found")

        tier = data.coverage_tier
        ee_cost = plan.get("employee_cost", {}).get(tier, 0)
        er_cost = plan.get("employer_contribution", {}).get(tier, 0)
        total = plan.get("premiums", {}).get(tier, 0)

        # Resolve add-ons
        add_ons = []
        for aid in data.add_on_plan_ids:
            ap = await db.plan_library.find_one({"id": aid}, {"_id": 0})
            if ap:
                add_ons.append({
                    "plan_id": ap["id"],
                    "plan_name": ap["plan_name"],
                    "category": ap.get("category", ""),
                    "employee_cost": ap.get("employee_cost", {}).get(tier, ap.get("employee_cost", {}).get("self_only", 0)),
                })

        enrollment = {
            "id": str(uuid.uuid4()),
            "employer_id": employer_id,
            "user_id": user["id"],
            "employee_id": user.get("linked_employee_id", ""),
            "employee_name": user["name"],
            "employee_email": user["email"],
            "plan_id": plan["id"],
            "plan_name": plan["plan_name"],
            "plan_type": plan["plan_type"],
            "carrier_name": plan.get("carrier_name", ""),
            "category": plan.get("category", "medical"),
            "coverage_tier": tier,
            "employee_premium": ee_cost,
            "employer_contribution": er_cost,
            "total_premium": total,
            "add_ons": add_ons,
            "status": "enrolled",
            "offer_code": OFFER_CODE_MAP.get(tier, "1E"),
            "enrolled_at": datetime.now(timezone.utc).isoformat(),
            "plan_year": plan.get("plan_year_start", "")[:4] or str(datetime.now(timezone.utc).year),
            "approved": False,
        }

        # Upsert enrollment
        await db.enrollments.update_one(
            {"employer_id": employer_id, "user_id": user["id"]},
            {"$set": enrollment},
            upsert=True,
        )

        # Update employee record if linked
        if user.get("linked_employee_id"):
            update = {
                "plan_id": plan["id"],
                "plan_name": plan["plan_name"],
                "coverage_tier": tier,
                "employee_monthly_premium": ee_cost,
                "employer_monthly_premium": er_cost,
                "offered_mec": True,
                "enrolled": True,
            }
            await db.employee_profiles.update_one({"id": user["linked_employee_id"]}, {"$set": update})
            await db.payroll_employees.update_one({"id": user["linked_employee_id"]}, {"$set": update})
            # Auto-update IRS offer code based on enrollment
            emp = await db.employee_profiles.find_one({"id": user["linked_employee_id"]}, {"_id": 0})
            if emp:
                await _recalculate_offer_code(db, employer_id, user["linked_employee_id"], emp)

        enrollment.pop("_id", None)
        result = await db.enrollments.find_one({"employer_id": employer_id, "user_id": user["id"]}, {"_id": 0})

        # Notify employer that employee enrolled
        employer_doc = await db.employers.find_one({"id": employer_id}, {"_id": 0, "user_id": 1})
        if employer_doc:
            await _create_notification(db, employer_doc["user_id"],
                "Employee Enrolled",
                f"{user.get('name', 'An employee')} has enrolled in {plan['plan_name']} ({tier.replace('_', ' ').title()}).",
                category="enrollment", link="/enrollment-review")

        return result

    @router.post("/enrollment/employee/decline")
    async def decline_employee(data: DeclineChoice, user=Depends(get_current_user)):
        """Employee declines coverage."""
        employer_id = user.get("employer_id")
        if not employer_id:
            raise HTTPException(status_code=400, detail="No employer linked")

        # Check enrollment window
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        active_period = await db.enrollment_periods.find_one(
            {"employer_id": employer_id, "status": "active"}, {"_id": 0}
        )
        if active_period:
            if today < active_period["start_date"] or today > active_period["end_date"]:
                exc = await db.enrollment_exceptions.find_one(
                    {"employer_id": employer_id, "employee_user_id": user["id"], "status": "approved"}, {"_id": 0}
                )
                if not exc:
                    raise HTTPException(status_code=403, detail="Enrollment window is closed. Request an exception from your employer.")
        if not active_period:
            any_period = await db.enrollment_periods.find_one({"employer_id": employer_id}, {"_id": 0})
            if any_period:
                exc = await db.enrollment_exceptions.find_one(
                    {"employer_id": employer_id, "employee_user_id": user["id"], "status": "approved"}, {"_id": 0}
                )
                if not exc:
                    raise HTTPException(status_code=403, detail="Enrollment window is closed. Request an exception from your employer.")

        # Calculate the offer code based on what was OFFERED, not enrollment status
        # The offer code should persist even when coverage is declined
        computed_offer_code = "1H"
        if user.get("linked_employee_id"):
            emp = await db.employee_profiles.find_one({"id": user["linked_employee_id"]}, {"_id": 0})
            if emp:
                computed_offer_code = await _recalculate_offer_code(db, employer_id, user["linked_employee_id"], emp)

        enrollment = {
            "id": str(uuid.uuid4()),
            "employer_id": employer_id,
            "user_id": user["id"],
            "employee_id": user.get("linked_employee_id", ""),
            "employee_name": user["name"],
            "employee_email": user["email"],
            "plan_id": "",
            "plan_name": "",
            "plan_type": "",
            "carrier_name": "",
            "category": "",
            "coverage_tier": "",
            "employee_premium": 0,
            "employer_contribution": 0,
            "total_premium": 0,
            "add_ons": [],
            "status": "declined",
            "decline_reason": data.reason,
            "decline_reason_detail": data.reason_detail,
            "offer_code": computed_offer_code,
            "enrolled_at": datetime.now(timezone.utc).isoformat(),
            "plan_year": str(datetime.now(timezone.utc).year),
            "approved": False,
        }

        await db.enrollments.update_one(
            {"employer_id": employer_id, "user_id": user["id"]},
            {"$set": enrollment},
            upsert=True,
        )

        # Update employee record - set offered_mec and enrolled status
        # Offer code is already set by _recalculate_offer_code above (reflects what was offered)
        if user.get("linked_employee_id"):
            await db.employee_profiles.update_one(
                {"id": user["linked_employee_id"]},
                {"$set": {"offered_mec": True, "enrolled": False}}
            )
            await db.payroll_employees.update_one(
                {"id": user["linked_employee_id"]},
                {"$set": {"offered_mec": True, "enrolled": False}}
            )

        result = await db.enrollments.find_one({"employer_id": employer_id, "user_id": user["id"]}, {"_id": 0})

        # Notify employer that employee declined
        employer_doc = await db.employers.find_one({"id": employer_id}, {"_id": 0, "user_id": 1})
        if employer_doc:
            reason_label = data.reason.replace("_", " ").title()
            await _create_notification(db, employer_doc["user_id"],
                "Employee Declined Coverage",
                f"{user.get('name', 'An employee')} has declined coverage. Reason: {reason_label}.",
                category="enrollment", link="/enrollment-review")

        return result

    # =============================================
    # STEP 4: HR COMPLIANCE REVIEW
    # =============================================

    @router.get("/enrollment/review/{employer_id}")
    async def get_enrollment_review(employer_id: str, user=Depends(get_current_user)):
        """HR views all enrollments for review."""
        enrollments = await db.enrollments.find(
            {"employer_id": employer_id}, {"_id": 0}
        ).sort("employee_name", 1).to_list(500)

        # Get eligibility results for context
        eligibility = await db.eligibility_results.find(
            {"employer_id": employer_id}, {"_id": 0}
        ).to_list(500)
        elig_map = {e["employee_id"]: e for e in eligibility}

        enrolled_count = sum(1 for e in enrollments if e["status"] == "enrolled")
        declined_count = sum(1 for e in enrollments if e["status"] == "declined")
        pending_count = sum(1 for e in enrollments if not e.get("approved"))

        # Decline reason breakdown
        decline_reasons = {}
        for e in enrollments:
            if e["status"] == "declined":
                reason = e.get("decline_reason", "other")
                decline_reasons[reason] = decline_reasons.get(reason, 0) + 1

        return {
            "total_enrollments": len(enrollments),
            "enrolled": enrolled_count,
            "declined": declined_count,
            "pending_approval": pending_count,
            "decline_reasons": decline_reasons,
            "enrollments": enrollments,
            "eligibility_map": elig_map,
        }

    @router.post("/enrollment/review/{employer_id}/approve-all")
    async def approve_all_enrollments(employer_id: str, user=Depends(get_current_user)):
        """HR approves all pending enrollments."""
        result = await db.enrollments.update_many(
            {"employer_id": employer_id, "approved": False},
            {"$set": {"approved": True, "approved_at": datetime.now(timezone.utc).isoformat(), "approved_by": user["id"]}}
        )
        return {"approved_count": result.modified_count}

    @router.put("/enrollment/review/{employer_id}/approve/{enrollment_id}")
    async def approve_enrollment(employer_id: str, enrollment_id: str, user=Depends(get_current_user)):
        """Approve a single enrollment."""
        result = await db.enrollments.update_one(
            {"id": enrollment_id, "employer_id": employer_id},
            {"$set": {"approved": True, "approved_at": datetime.now(timezone.utc).isoformat(), "approved_by": user["id"]}}
        )
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Enrollment not found")
        return {"message": "Approved"}

    @router.get("/enrollment/review/{employer_id}/proof/{enrollment_id}")
    async def download_enrollment_proof(employer_id: str, enrollment_id: str, user=Depends(get_current_user)):
        """Generate a PDF proof of enrollment or coverage decline."""
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch

        enrollment = await db.enrollments.find_one(
            {"id": enrollment_id, "employer_id": employer_id}, {"_id": 0}
        )
        if not enrollment:
            raise HTTPException(status_code=404, detail="Enrollment not found")

        employer = await db.employers.find_one({"id": employer_id}, {"_id": 0})
        employer_name = employer.get("company_name", "Employer") if employer else "Employer"

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.6*inch, bottomMargin=0.6*inch, leftMargin=0.7*inch, rightMargin=0.7*inch)
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('DocTitle', parent=styles['Title'], fontSize=16, spaceAfter=4, textColor=colors.HexColor('#1e293b'))
        subtitle_style = ParagraphStyle('SubTitle', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#64748b'), spaceAfter=14)
        section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=11, textColor=colors.HexColor('#334155'), spaceBefore=14, spaceAfter=6)
        normal_style = ParagraphStyle('NormalText', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#334155'), spaceAfter=3)
        small_style = ParagraphStyle('SmallText', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#94a3b8'))

        is_enrolled = enrollment.get("status") == "enrolled"
        story = []

        # Title
        if is_enrolled:
            story.append(Paragraph("Certificate of Coverage Enrollment", title_style))
            story.append(Paragraph("Proof that the employee has elected health coverage under the employer-sponsored plan.", subtitle_style))
        else:
            story.append(Paragraph("Certificate of Coverage Decline", title_style))
            story.append(Paragraph("Proof that the employee has declined health coverage under the employer-sponsored plan.", subtitle_style))

        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e2e8f0')))
        story.append(Spacer(1, 10))

        # Employee & Employer Info
        story.append(Paragraph("Employee Information", section_style))
        emp_data = [
            ["Employee Name", enrollment.get("employee_name", "—")],
            ["Email", enrollment.get("employee_email", "—")],
            ["Employer", employer_name],
            ["Plan Year", enrollment.get("plan_year", str(datetime.now(timezone.utc).year))],
            ["Decision Date", enrollment.get("enrolled_at", "—")[:10] if enrollment.get("enrolled_at") else "—"],
        ]
        t = Table(emp_data, colWidths=[2*inch, 4.4*inch])
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#64748b')),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1e293b')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('LINEBELOW', (0, 0), (-1, -2), 0.5, colors.HexColor('#f1f5f9')),
        ]))
        story.append(t)
        story.append(Spacer(1, 6))

        # Coverage Decision
        story.append(Paragraph("Coverage Decision", section_style))
        if is_enrolled:
            story.append(Paragraph(f"<b>Status:</b> ENROLLED — Coverage Accepted", normal_style))
            plan_data = [
                ["Plan Name", enrollment.get("plan_name", "—")],
                ["Carrier", enrollment.get("carrier_name", "—")],
                ["Plan Type", enrollment.get("plan_type", "—")],
                ["Category", (enrollment.get("category", "—") or "—").title()],
                ["Coverage Tier", (enrollment.get("coverage_tier", "—") or "—").replace("_", " ").title()],
                ["Offer Code", enrollment.get("offer_code", "—")],
            ]
            t2 = Table(plan_data, colWidths=[2*inch, 4.4*inch])
            t2.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#64748b')),
                ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1e293b')),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('LINEBELOW', (0, 0), (-1, -2), 0.5, colors.HexColor('#f1f5f9')),
            ]))
            story.append(t2)
            story.append(Spacer(1, 6))

            # Premium Details
            story.append(Paragraph("Premium Details", section_style))
            premium_data = [
                ["Employee Monthly Premium", f"${enrollment.get('employee_premium', 0):,.2f}"],
                ["Employer Contribution", f"${enrollment.get('employer_contribution', 0):,.2f}"],
                ["Total Monthly Premium", f"${enrollment.get('total_premium', 0):,.2f}"],
            ]
            t3 = Table(premium_data, colWidths=[2*inch, 4.4*inch])
            t3.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#64748b')),
                ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1e293b')),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('LINEBELOW', (0, 0), (-1, -2), 0.5, colors.HexColor('#f1f5f9')),
            ]))
            story.append(t3)

            # Add-ons
            if enrollment.get("add_ons"):
                story.append(Spacer(1, 6))
                story.append(Paragraph("Additional Coverage (Add-Ons)", section_style))
                for addon in enrollment["add_ons"]:
                    story.append(Paragraph(f"• {addon.get('plan_name', '—')} ({addon.get('category', '').title()}) — ${addon.get('employee_cost', 0):,.2f}/mo", normal_style))
        else:
            story.append(Paragraph(f"<b>Status:</b> DECLINED — Coverage Waived", normal_style))
            decline_reason = (enrollment.get("decline_reason", "—") or "—").replace("_", " ").title()
            decline_detail = enrollment.get("decline_reason_detail", "")
            decline_data = [
                ["Decline Reason", decline_reason],
            ]
            if decline_detail:
                decline_data.append(["Additional Details", decline_detail])
            decline_data.append(["Offer Code", enrollment.get("offer_code", "—")])

            t2 = Table(decline_data, colWidths=[2*inch, 4.4*inch])
            t2.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#64748b')),
                ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1e293b')),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('LINEBELOW', (0, 0), (-1, -2), 0.5, colors.HexColor('#f1f5f9')),
            ]))
            story.append(t2)

        # Approval Status
        story.append(Spacer(1, 10))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e2e8f0')))
        story.append(Spacer(1, 6))
        story.append(Paragraph("Approval Status", section_style))
        if enrollment.get("approved"):
            story.append(Paragraph(f"<b>HR Approved:</b> Yes — {enrollment.get('approved_at', '')[:10] if enrollment.get('approved_at') else ''}", normal_style))
        else:
            story.append(Paragraph("<b>HR Approved:</b> Pending", normal_style))

        # Footer
        story.append(Spacer(1, 20))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#e2e8f0')))
        story.append(Spacer(1, 6))
        story.append(Paragraph(f"Generated on {datetime.now(timezone.utc).strftime('%B %d, %Y at %H:%M UTC')} — {employer_name} ACA Compliance System", small_style))
        story.append(Paragraph("This document serves as official proof of the employee's coverage election or waiver decision.", small_style))

        doc.build(story)
        buf.seek(0)

        safe_name = enrollment.get("employee_name", "employee").replace(" ", "_")
        status_tag = "enrolled" if is_enrolled else "declined"
        filename = f"{safe_name}_{status_tag}_proof.pdf"

        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    # =============================================
    # STEP 4B: PAYROLL DEDUCTION FILE (ADP/UKG)
    # =============================================

    @router.get("/enrollment/payroll-export/{employer_id}")
    async def generate_payroll_deduction_file(employer_id: str, user=Depends(get_current_user)):
        """Generate ADP/UKG-compatible payroll deduction file from enrolled employees."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        enrollments = await db.enrollments.find(
            {"employer_id": employer_id, "status": "enrolled"}, {"_id": 0}
        ).sort("employee_name", 1).to_list(500)

        if not enrollments:
            raise HTTPException(status_code=400, detail="No enrolled employees to export")

        employer = await db.employers.find_one({"id": employer_id}, {"_id": 0})
        employer_name = employer.get("company_name", "Employer") if employer else "Employer"

        wb = Workbook()
        ws = wb.active
        ws.title = "Payroll Deductions"

        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="B91C1C", end_color="B91C1C", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        headers = [
            "Employee ID", "Employee Name", "Email", "Plan Name", "Carrier",
            "Coverage Tier", "Deduction Type", "Deduction Amount (Monthly)",
            "Employer Contribution (Monthly)", "Total Premium", "Effective Date",
            "Frequency", "Pre-Tax"
        ]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for i, e in enumerate(enrollments, 2):
            emp = None
            if e.get("employee_id"):
                emp = await db.employee_profiles.find_one({"id": e["employee_id"]}, {"_id": 0})
            emp_id = emp.get("employee_id", e.get("employee_id", "")) if emp else e.get("employee_id", "")

            ws.cell(row=i, column=1, value=emp_id).border = thin_border
            ws.cell(row=i, column=2, value=e["employee_name"]).border = thin_border
            ws.cell(row=i, column=3, value=e.get("employee_email", "")).border = thin_border
            ws.cell(row=i, column=4, value=e["plan_name"]).border = thin_border
            ws.cell(row=i, column=5, value=e.get("carrier_name", "")).border = thin_border
            ws.cell(row=i, column=6, value=e["coverage_tier"].replace("_", " ").title()).border = thin_border
            ws.cell(row=i, column=7, value="Health Insurance").border = thin_border
            c = ws.cell(row=i, column=8, value=e["employee_premium"])
            c.number_format = '$#,##0.00'
            c.border = thin_border
            c = ws.cell(row=i, column=9, value=e["employer_contribution"])
            c.number_format = '$#,##0.00'
            c.border = thin_border
            c = ws.cell(row=i, column=10, value=e["total_premium"])
            c.number_format = '$#,##0.00'
            c.border = thin_border
            ws.cell(row=i, column=11, value=e.get("enrolled_at", "")[:10]).border = thin_border
            ws.cell(row=i, column=12, value="Monthly").border = thin_border
            ws.cell(row=i, column=13, value="Yes").border = thin_border

            # Add-on rows
            for addon in e.get("add_ons", []):
                i += 1
                ws.cell(row=i, column=1, value=emp_id).border = thin_border
                ws.cell(row=i, column=2, value=e["employee_name"]).border = thin_border
                ws.cell(row=i, column=3, value=e.get("employee_email", "")).border = thin_border
                ws.cell(row=i, column=4, value=addon.get("plan_name", "")).border = thin_border
                ws.cell(row=i, column=5, value="").border = thin_border
                ws.cell(row=i, column=6, value=e["coverage_tier"].replace("_", " ").title()).border = thin_border
                ws.cell(row=i, column=7, value=addon.get("category", "Supplemental").title()).border = thin_border
                c = ws.cell(row=i, column=8, value=addon.get("employee_cost", 0))
                c.number_format = '$#,##0.00'
                c.border = thin_border
                ws.cell(row=i, column=9, value=0).border = thin_border
                ws.cell(row=i, column=10, value=addon.get("employee_cost", 0)).border = thin_border
                ws.cell(row=i, column=11, value=e.get("enrolled_at", "")[:10]).border = thin_border
                ws.cell(row=i, column=12, value="Monthly").border = thin_border
                ws.cell(row=i, column=13, value="Yes").border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = max_len

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"payroll_deductions_{employer_name.replace(' ', '_')}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    @router.get("/enrollment/payroll-summary/{employer_id}")
    async def get_payroll_deduction_summary(employer_id: str, user=Depends(get_current_user)):
        """Get summary of deductions ready to send to payroll."""
        enrollments = await db.enrollments.find(
            {"employer_id": employer_id, "status": "enrolled"}, {"_id": 0}
        ).to_list(500)

        total_ee_deductions = sum(e.get("employee_premium", 0) for e in enrollments)
        total_er_contributions = sum(e.get("employer_contribution", 0) for e in enrollments)
        carriers = list(set(e.get("carrier_name", "Unknown") for e in enrollments if e.get("carrier_name")))

        return {
            "enrolled_count": len(enrollments),
            "total_ee_deductions": round(total_ee_deductions, 2),
            "total_er_contributions": round(total_er_contributions, 2),
            "total_premium": round(total_ee_deductions + total_er_contributions, 2),
            "carriers": carriers,
            "ready": len(enrollments) > 0,
        }

    @router.get("/enrollment/carrier-export/{employer_id}")
    async def generate_carrier_census(employer_id: str, carrier: str = "", user=Depends(get_current_user)):
        """Generate insurance carrier census Excel for enrolled employees."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        query = {"employer_id": employer_id, "status": "enrolled"}
        if carrier:
            query["carrier_name"] = carrier

        enrollments = await db.enrollments.find(query, {"_id": 0}).sort("employee_name", 1).to_list(500)

        if not enrollments:
            raise HTTPException(status_code=400, detail="No enrolled employees found")

        employer = await db.employers.find_one({"id": employer_id}, {"_id": 0})
        employer_name = employer.get("company_name", "Employer") if employer else "Employer"

        wb = Workbook()
        ws = wb.active
        ws.title = "Carrier Census"

        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        headers = [
            "SSN (Last 4)", "Employee Name", "Date of Birth", "Plan Name",
            "Coverage Tier", "Effective Date", "Employee Premium",
            "Employer Contribution", "Total Premium", "Dependents"
        ]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for i, e in enumerate(enrollments, 2):
            emp = None
            if e.get("employee_id"):
                emp = await db.employee_profiles.find_one({"id": e["employee_id"]}, {"_id": 0})

            ssn = emp.get("ssn_last4", "XXXX") if emp else "XXXX"
            dob = emp.get("date_of_birth", "") if emp else ""
            deps = len(emp.get("dependents", [])) if emp else 0

            ws.cell(row=i, column=1, value=f"***-**-{ssn}").border = thin_border
            ws.cell(row=i, column=2, value=e["employee_name"]).border = thin_border
            ws.cell(row=i, column=3, value=dob).border = thin_border
            ws.cell(row=i, column=4, value=e["plan_name"]).border = thin_border
            ws.cell(row=i, column=5, value=e["coverage_tier"].replace("_", " ").title()).border = thin_border
            ws.cell(row=i, column=6, value=e.get("enrolled_at", "")[:10]).border = thin_border
            c = ws.cell(row=i, column=7, value=e["employee_premium"])
            c.number_format = '$#,##0.00'
            c.border = thin_border
            c = ws.cell(row=i, column=8, value=e["employer_contribution"])
            c.number_format = '$#,##0.00'
            c.border = thin_border
            c = ws.cell(row=i, column=9, value=e["total_premium"])
            c.number_format = '$#,##0.00'
            c.border = thin_border
            ws.cell(row=i, column=10, value=deps).border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = max_len

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        carrier_tag = carrier.replace(" ", "_") if carrier else "all_carriers"
        filename = f"carrier_census_{carrier_tag}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    # =============================================
    # STEP 5: CARRIER CENSUS EXPORT
    # =============================================

    @router.post("/enrollment/census/{employer_id}")
    async def generate_census(employer_id: str, user=Depends(get_current_user)):
        """Generate census data for carrier submission."""
        enrollments = await db.enrollments.find(
            {"employer_id": employer_id, "status": "enrolled", "approved": True}, {"_id": 0}
        ).to_list(500)

        if not enrollments:
            raise HTTPException(status_code=400, detail="No approved enrollments found. Approve enrollments first.")

        employer = await db.employers.find_one({"id": employer_id}, {"_id": 0})

        # Build census rows
        census_rows = []
        for e in enrollments:
            # Try to get SSN from employee profile
            emp = None
            if e.get("employee_id"):
                emp = await db.employee_profiles.find_one({"id": e["employee_id"]}, {"_id": 0})
                if not emp:
                    emp = await db.payroll_employees.find_one({"id": e["employee_id"]}, {"_id": 0})

            census_rows.append({
                "ssn_last4": emp.get("ssn_last4", "XXXX") if emp else "XXXX",
                "employee_name": e["employee_name"],
                "plan_name": e["plan_name"],
                "carrier": e.get("carrier_name", ""),
                "coverage_tier": e["coverage_tier"].replace("_", " ").title(),
                "employee_premium": e["employee_premium"],
                "employer_contribution": e["employer_contribution"],
                "total_premium": e["total_premium"],
                "effective_date": e.get("enrolled_at", "")[:10],
                "offer_code": e.get("offer_code", ""),
            })

        # Save census
        census_id = str(uuid.uuid4())
        census_doc = {
            "id": census_id,
            "employer_id": employer_id,
            "employer_name": employer.get("name", "") if employer else "",
            "total_enrolled": len(census_rows),
            "rows": census_rows,
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.census_exports.insert_one(census_doc)
        census_doc.pop("_id", None)
        return census_doc

    @router.get("/enrollment/census/{employer_id}/download/{census_id}")
    async def download_census_excel(employer_id: str, census_id: str, user=Depends(get_current_user)):
        """Download census as Excel file."""
        census = await db.census_exports.find_one({"id": census_id, "employer_id": employer_id}, {"_id": 0})
        if not census:
            raise HTTPException(status_code=404, detail="Census not found")

        wb = Workbook()
        ws = wb.active
        ws.title = "Carrier Census"

        # Header style
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        headers = ["SSN (Last 4)", "Employee Name", "Plan Name", "Carrier", "Coverage Tier",
                    "Employee Premium", "Employer Contribution", "Total Premium", "Effective Date", "Offer Code"]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for i, row in enumerate(census["rows"], 2):
            ws.cell(row=i, column=1, value=row["ssn_last4"]).border = thin_border
            ws.cell(row=i, column=2, value=row["employee_name"]).border = thin_border
            ws.cell(row=i, column=3, value=row["plan_name"]).border = thin_border
            ws.cell(row=i, column=4, value=row["carrier"]).border = thin_border
            ws.cell(row=i, column=5, value=row["coverage_tier"]).border = thin_border
            c = ws.cell(row=i, column=6, value=row["employee_premium"])
            c.number_format = '$#,##0.00'
            c.border = thin_border
            c = ws.cell(row=i, column=7, value=row["employer_contribution"])
            c.number_format = '$#,##0.00'
            c.border = thin_border
            c = ws.cell(row=i, column=8, value=row["total_premium"])
            c.number_format = '$#,##0.00'
            c.border = thin_border
            ws.cell(row=i, column=9, value=row["effective_date"]).border = thin_border
            ws.cell(row=i, column=10, value=row["offer_code"]).border = thin_border

        # Auto-width columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = max_len

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"census_{census['employer_name'].replace(' ', '_')}_{census['generated_at'][:10]}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    @router.get("/enrollment/census-history/{employer_id}")
    async def get_census_history(employer_id: str, user=Depends(get_current_user)):
        """Get census export history."""
        history = await db.census_exports.find(
            {"employer_id": employer_id}, {"_id": 0, "rows": 0}
        ).sort("generated_at", -1).to_list(20)
        return history

    # =============================================
    # OPEN ENROLLMENT PERIOD MANAGEMENT
    # =============================================

    @router.post("/enrollment/periods")
    async def create_enrollment_period(data: dict, user=Depends(get_current_user)):
        """HR creates an enrollment window."""
        employer_id = data.get("employer_id")
        if not employer_id:
            raise HTTPException(status_code=400, detail="employer_id required")
        period = {
            "id": str(uuid.uuid4()),
            "employer_id": employer_id,
            "period_name": data.get("period_name", "Open Enrollment"),
            "start_date": data["start_date"],
            "end_date": data["end_date"],
            "status": "draft",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.enrollment_periods.insert_one(period)
        period.pop("_id", None)
        return period

    @router.get("/enrollment/periods/{employer_id}")
    async def get_enrollment_periods(employer_id: str, user=Depends(get_current_user)):
        """Get all enrollment periods for employer."""
        periods = await db.enrollment_periods.find(
            {"employer_id": employer_id}, {"_id": 0}
        ).sort("created_at", -1).to_list(50)
        # Auto-close expired active periods
        now = datetime.now(timezone.utc).isoformat()
        for p in periods:
            if p["status"] == "active" and p["end_date"] < now[:10]:
                await db.enrollment_periods.update_one(
                    {"id": p["id"]}, {"$set": {"status": "closed", "updated_at": now}}
                )
                p["status"] = "closed"
        return periods

    @router.put("/enrollment/periods/{period_id}")
    async def update_enrollment_period(period_id: str, data: dict, user=Depends(get_current_user)):
        """Update enrollment period (activate, close, edit dates)."""
        update = {"updated_at": datetime.now(timezone.utc).isoformat()}
        for f in ["period_name", "start_date", "end_date", "status"]:
            if f in data:
                update[f] = data[f]
        # If activating, deactivate other active periods for this employer
        if data.get("status") == "active":
            period = await db.enrollment_periods.find_one({"id": period_id}, {"_id": 0})
            if period:
                await db.enrollment_periods.update_many(
                    {"employer_id": period["employer_id"], "status": "active", "id": {"$ne": period_id}},
                    {"$set": {"status": "closed", "updated_at": datetime.now(timezone.utc).isoformat()}}
                )
        result = await db.enrollment_periods.update_one({"id": period_id}, {"$set": update})
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Period not found")
        updated = await db.enrollment_periods.find_one({"id": period_id}, {"_id": 0})
        return updated

    @router.delete("/enrollment/periods/{period_id}")
    async def delete_enrollment_period(period_id: str, user=Depends(get_current_user)):
        """Delete a draft enrollment period."""
        result = await db.enrollment_periods.delete_one({"id": period_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Period not found")
        return {"message": "Deleted"}

    @router.get("/enrollment/periods/{employer_id}/active")
    async def get_active_enrollment_period(employer_id: str, user=Depends(get_current_user)):
        """Get the current enrollment period (active or recently closed) for employee portal."""
        now = datetime.now(timezone.utc)
        today = now.strftime("%Y-%m-%d")
        # Look for active period first
        period = await db.enrollment_periods.find_one(
            {"employer_id": employer_id, "status": "active"}, {"_id": 0}
        )
        if period:
            if period["end_date"] < today:
                await db.enrollment_periods.update_one(
                    {"id": period["id"]},
                    {"$set": {"status": "closed", "updated_at": now.isoformat()}}
                )
                period["status"] = "closed"
        # If no active period, find the most recently closed one
        if not period:
            period = await db.enrollment_periods.find_one(
                {"employer_id": employer_id, "status": "closed"}, {"_id": 0},
                sort=[("end_date", -1)]
            )
        # Check if employee has an exception (any status)
        has_exception = False
        exception_status = None
        if user.get("role") == "employee" and period:
            exc = await db.enrollment_exceptions.find_one(
                {"employer_id": employer_id, "employee_user_id": user["id"]},
                {"_id": 0},
                sort=[("created_at", -1)]
            )
            if exc:
                exception_status = exc.get("status")  # pending, approved, rejected
                has_exception = exception_status == "approved"
        return {"period": period, "has_exception": has_exception, "exception_status": exception_status}

    # --- Exception Requests ---

    @router.post("/enrollment/exceptions")
    async def request_enrollment_exception(data: dict, user=Depends(get_current_user)):
        """Employee requests an exception to enroll outside the open window."""
        employer_id = user.get("employer_id")
        if not employer_id:
            raise HTTPException(status_code=400, detail="No employer linked")
        existing = await db.enrollment_exceptions.find_one(
            {"employer_id": employer_id, "employee_user_id": user["id"], "status": "pending"},
            {"_id": 0}
        )
        if existing:
            raise HTTPException(status_code=400, detail="You already have a pending exception request")
        exc = {
            "id": str(uuid.uuid4()),
            "employer_id": employer_id,
            "employee_user_id": user["id"],
            "employee_name": user.get("name", ""),
            "employee_email": user.get("email", ""),
            "reason": data.get("reason", ""),
            "status": "pending",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.enrollment_exceptions.insert_one(exc)
        exc.pop("_id", None)

        # Notify employer about exception request
        employer_doc = await db.employers.find_one({"id": employer_id}, {"_id": 0, "user_id": 1})
        if employer_doc:
            await _create_notification(db, employer_doc["user_id"],
                "Enrollment Exception Request",
                f"{user.get('name', 'An employee')} has requested an enrollment exception: {data.get('reason', '')}",
                category="exception", link="/enrollment-review")

        return exc

    @router.get("/enrollment/exceptions/{employer_id}")
    async def get_enrollment_exceptions(employer_id: str, user=Depends(get_current_user)):
        """HR views all exception requests."""
        exceptions = await db.enrollment_exceptions.find(
            {"employer_id": employer_id}, {"_id": 0}
        ).sort("created_at", -1).to_list(100)
        return exceptions

    @router.put("/enrollment/exceptions/{exception_id}")
    async def review_enrollment_exception(exception_id: str, data: dict, user=Depends(get_current_user)):
        """HR approves or rejects an exception."""
        new_status = data.get("status")
        if new_status not in ("approved", "rejected"):
            raise HTTPException(status_code=400, detail="Status must be approved or rejected")
        result = await db.enrollment_exceptions.update_one(
            {"id": exception_id},
            {"$set": {
                "status": new_status,
                "reviewed_at": datetime.now(timezone.utc).isoformat(),
                "reviewed_by": user["id"],
            }}
        )
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Exception not found")
        updated = await db.enrollment_exceptions.find_one({"id": exception_id}, {"_id": 0})

        # Notify employee about HR decision
        if updated:
            emp_user_id = updated.get("employee_user_id")
            if emp_user_id:
                status_word = "approved" if new_status == "approved" else "rejected"
                msg = (
                    "Your enrollment exception request has been approved. You can now enroll in your assigned plan."
                    if new_status == "approved"
                    else "Your enrollment exception request has been rejected. Please contact your HR administrator for more information."
                )
                await _create_notification(db, emp_user_id,
                    f"Exception Request {status_word.title()}",
                    msg,
                    category="exception", link="/employee-portal")

        return updated

